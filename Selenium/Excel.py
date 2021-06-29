import openpyxl

f=openpyxl.load_workbook(r'C:\Users\umasankar.panigrahy\Desktop\Docker\PM\test_data\employee.xlsx')


sheet=f.get_sheet_by_name('Sheet1')
total_row=(sheet.max_row)
print(total_row)                  #row count
total_col= (sheet.max_column)
print(total_col)                  #column count

print("*****************")

active=f.active

row1=active.cell(row=1, column=1)
print(row1.value)                   #value attribute we can can print the value of cell
print("*****************")

#print all data
read_row= active['A1': 'C10']
for i1,i2,i3 in read_row:
    print("{0:10}{1:10}{2:10}".format(i1.value, i2.value,i3.value))
    if i1.value=='Kanhu':
        break
print("*****************")

#Print a particular row value
for p in range (1, total_col+1):
    specific_row=active.cell(row=2, column=p)
    print(specific_row.value,  end=" ")
print("\n*****************")

# print all column
for i in range(1, total_col+1):
    all_column=active.cell(row=1,column=i)
    print(all_column.value)
print("*****************")

#Print first column value
for r in range (1,total_row+1):
    first_column=active.cell(row=r, column=1)
    print(first_column.value)
print("*****************")

"""
print(f.sheetnames)         #return all sheetnames

print('active sheet is', f.active.title)    #return active sheets

sheet=f['Sheet1']
print(sheet['A2'].row)
a=sheet.cell(2,2)
print(a.value)
"""